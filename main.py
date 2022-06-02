import os
import openpyxl
from openpyxl import load_workbook

root_dir = os.getcwd()
xlsx_dir = os.path.join(root_dir, 'xlsx')
result_dir = os.path.join(root_dir, 'result')

if not (os.path.exists(xlsx_dir)):
    os.makedirs(xlsx_dir)

if not (os.path.exists(result_dir)):
    os.makedirs(result_dir)

if(__name__ == "__main__"):
    print('Программа запущена!!!')

    for file_one in os.listdir(xlsx_dir):
        file_path = os.path.join(xlsx_dir, file_one)
        file_one_name = file_one.split('.')[0]

        wb = load_workbook(filename=file_path)
        ws = wb.active
        max_row = ws.max_row
        max_col = ws.max_column

        # Создаём пустые словари
        col_data = dict()
        for i in range(1, max_col + 1):
            n = file_one_name + '_n' + str(i)
            col_data[n] = []

        # Заполняем пустые словари данными
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                value = ws.cell(column=col, row=row).value
                if value:
                    n = file_one_name + '_n' + str(col)
                    col_data[n].append(value)

        # Создаём файл и кладём в него данные столбца
        for i in range(1, max_col + 1):
            n = file_one_name + '_n' + str(i)
            new_file_name = n + '.xlsx'
            new_file_path = os.path.join(result_dir, new_file_name)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Лист1'

            nu = 1
            for m in col_data[n]:
                cell_coord_A = "A" + str(nu)
                a = ws[cell_coord_A]
                a.value = m
                nu += 1

            wb.save(new_file_path)

    print('Программа выполнена!!!')
    print('')
    print('***')
    print('Название программы: ColsToFiles v1.0')
    print('Разработка: Министерства труда и социального развития РД')
    print('Разработчик: Ахмедов Мурад Алилович')
    print('***')
    print('')
    input('Нажмите любую клавишу, для завершения работы программы')